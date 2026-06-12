#!/usr/bin/env python3
"""
create_manifest_xlsx.py
Creates a styled Excel manifest template for a given sprint.

Usage:
    python3 scripts/create_manifest_xlsx.py sprint-12
"""

import sys
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Error: openpyxl not found. Install with:  pip3 install openpyxl")


# ── Colour constants ──────────────────────────────────────────────────────────
DARK_BLUE   = "1A233A"
WHITE       = "FFFFFF"
ROW_ALT     = "DEF0FA"   # light blue alternating rows
GIT_GREEN   = "D9EAD3"
SP_AMBER    = "FFF2CC"


def make_header_fill():
    return PatternFill("solid", fgColor=DARK_BLUE)

def make_alt_fill():
    return PatternFill("solid", fgColor=ROW_ALT)

def make_white_fill():
    return PatternFill("solid", fgColor=WHITE)

def header_font():
    return Font(color=WHITE, bold=True)

def set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


# ── Sheet 1: Build Info ───────────────────────────────────────────────────────
def build_info_sheet(wb, sprint):
    ws = wb.create_sheet("Build Info")

    headers = ["Field", "Value"]
    rows = [
        ["release",        sprint],
        ["sprint_dates",   "YYYY-MM-DD → YYYY-MM-DD"],
        ["build_date",     ""],
        ["built_by",       ""],
        ["elm_sprint_url", ""],
    ]

    # Header row
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = make_header_fill()
        cell.font = header_font()
        cell.alignment = Alignment(horizontal="left")

    # Data rows with alternating fill
    for r_idx, row_data in enumerate(rows, 2):
        fill = make_white_fill() if (r_idx % 2 == 0) else make_alt_fill()
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill

    # Column widths
    set_col_width(ws, 1, 22)
    set_col_width(ws, 2, 45)


# ── Sheet 2: Teams ────────────────────────────────────────────────────────────
def teams_sheet(wb, sprint):
    ws = wb.create_sheet("Teams")

    headers = [
        "team_key", "team_name", "source_control", "repo_or_sharepoint_url",
        "branch", "commit_or_artifact_version", "included", "elm_work_items",
        "summary", "notes", "depends_on", "reason_if_deferred",
    ]

    col_widths = [18, 28, 14, 55, 20, 22, 10, 28, 35, 25, 22, 35]

    sp_url_base  = f"https://company.sharepoint.com/sites/ETL/Shared Documents/{sprint}"
    git_branch   = f"sprint/{sprint}"

    data_rows = [
        ["conversion",      "Conversion Team",           "sharepoint", sp_url_base,
         "N/A", "", "TRUE", "", "", "", "", ""],
        ["interfaces",      "Interfaces Team",           "git",        f"satishreddy13/repo-interfaces",
         git_branch, "", "TRUE", "", "", "", "", ""],
        ["workflow_config",  "Workflow Config Team",      "git",        f"satishreddy13/repo-workflow-config",
         git_branch, "", "TRUE", "", "", "", "", ""],
        ["func_config",     "Functional Product Config", "git",        f"satishreddy13/repo-func-config",
         git_branch, "", "TRUE", "", "", "", "", ""],
    ]

    # Header row
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = make_header_fill()
        cell.font = header_font()
        cell.alignment = Alignment(horizontal="left")

    # Data rows
    for r_idx, row_data in enumerate(data_rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            # Colour the source_control cell (col 3) based on type
            if c_idx == 3:
                sc = val
                if sc == "git":
                    cell.fill = PatternFill("solid", fgColor=GIT_GREEN)
                elif sc == "sharepoint":
                    cell.fill = PatternFill("solid", fgColor=SP_AMBER)

    # Column widths
    for col_idx, width in enumerate(col_widths, 1):
        set_col_width(ws, col_idx, width)

    # Freeze row 1
    ws.freeze_panes = "A2"


# ── Sheet 3: COTS ─────────────────────────────────────────────────────────────
def cots_sheet(wb, sprint):
    ws = wb.create_sheet("COTS")

    headers = ["Field", "Value"]
    rows = [
        ["product",            "Acme ETL Platform"],
        ["version",            ""],
        ["previous_version",   ""],
        ["hotfixes_included",  ""],
        ["included",           "TRUE"],
        ["notes",              ""],
        ["release_notes_file", f"cots-{sprint}.md"],
    ]

    # Header row
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = make_header_fill()
        cell.font = header_font()
        cell.alignment = Alignment(horizontal="left")

    # Data rows with alternating fill
    for r_idx, row_data in enumerate(rows, 2):
        fill = make_white_fill() if (r_idx % 2 == 0) else make_alt_fill()
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = fill

    set_col_width(ws, 1, 22)
    set_col_width(ws, 2, 45)


# ── Sheet 4: Environments ─────────────────────────────────────────────────────
def environments_sheet(wb):
    ws = wb.create_sheet("Environments")

    headers = ["environment", "status", "deployed_by", "deployed_date", "passed_date", "notes"]
    col_widths = [14, 12, 18, 18, 18, 35]

    data_rows = [
        ["sit", "pending", "", "", "", ""],
        ["uat", "pending", "", "", "", ""],
    ]

    # Header row
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = make_header_fill()
        cell.font = header_font()
        cell.alignment = Alignment(horizontal="left")

    # Data rows
    for r_idx, row_data in enumerate(data_rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    for col_idx, width in enumerate(col_widths, 1):
        set_col_width(ws, col_idx, width)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: create_manifest_xlsx.py <sprint>   e.g. create_manifest_xlsx.py sprint-12")

    sprint = sys.argv[1]

    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root  = os.path.dirname(script_dir)
    sprint_dir = os.path.join(repo_root, "releases", sprint)
    os.makedirs(sprint_dir, exist_ok=True)

    out_path = os.path.join(sprint_dir, "manifest.xlsx")

    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    build_info_sheet(wb, sprint)
    teams_sheet(wb, sprint)
    cots_sheet(wb, sprint)
    environments_sheet(wb)

    wb.save(out_path)
    print(f"Created: releases/{sprint}/manifest.xlsx")


if __name__ == "__main__":
    main()
