#!/usr/bin/env python3
"""
generate-report.py
Reads releases/<sprint>/manifest.xlsx and writes releases/<sprint>/RELEASE-NOTES.xlsx.

Usage:
    python3 scripts/generate-report.py sprint-12
"""

import sys
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Error: openpyxl not found. Install with:  pip3 install openpyxl")


# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BLUE  = "1A233A"
MID_BLUE   = "1F4E79"
ACCENT     = "2E75B6"
GREEN_DARK = "375F1B"
GREEN_FILL = "E2EFDA"
AMBER_FILL = "FFF2CC"
RED_FILL   = "FCE4D6"
GREY_FILL  = "F2F2F2"
LIGHT_BLUE = "DEEAF1"
WHITE      = "FFFFFF"

def hfill(hex_str):
    return PatternFill("solid", fgColor=hex_str)

def hfont(hex_str, bold=False, size=11, name="Calibri"):
    return Font(color=hex_str, bold=bold, size=size, name=name)

def border_bottom(hex_str="BFBFBF"):
    s = Side(style="thin", color=hex_str)
    return Border(bottom=s)

def thick_border():
    s = Side(style="medium", color="1F4E79")
    return Border(bottom=s)

def set_col_widths(ws, widths):
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ── Parsing helpers ───────────────────────────────────────────────────────────

def cell_str(cell):
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()

def parse_kv_sheet(ws):
    result = {}
    for row in ws.iter_rows(min_row=2):
        key = cell_str(row[0])
        val = cell_str(row[1]) if len(row) > 1 else ""
        if key:
            result[key] = val
    return result

def parse_bool(val):
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().upper() == "TRUE"
    return bool(val)

def parse_teams_sheet(ws):
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [cell_str(c) for c in header_row]
    teams = {}
    for row in ws.iter_rows(min_row=2):
        values = [cell_str(c) for c in row]
        if not any(values):
            continue
        record = dict(zip(headers, values))
        included_raw = row[headers.index("included")].value if "included" in headers else ""
        record["included"] = parse_bool(included_raw)
        key = record.get("team_key", "")
        if key:
            teams[key] = record
    return teams

def parse_environments_sheet(ws):
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [cell_str(c) for c in header_row]
    envs = {}
    for row in ws.iter_rows(min_row=2):
        values = [cell_str(c) for c in row]
        if not any(values):
            continue
        record = dict(zip(headers, values))
        env_name = record.get("environment", "")
        if env_name:
            envs[env_name] = record
    return envs


TEAM_ORDER = ["conversion", "interfaces", "workflow_config", "func_config"]
TEAM_LABELS = {
    "conversion":      "Conversion",
    "interfaces":      "Interfaces",
    "workflow_config": "Workflow Config",
    "func_config":     "Functional Product Config",
}


# ── Sheet writers ─────────────────────────────────────────────────────────────

def write_summary_sheet(wb, sprint, build_info, envs, generated_at):
    ws = wb.create_sheet("Summary")

    # ── Title bar ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"Release Notes — {build_info.get('release', sprint)}"
    c.fill  = hfill(DARK_BLUE)
    c.font  = hfont(WHITE, bold=True, size=14)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # ── Build info header ─────────────────────────────────────────────────────
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = "BUILD INFORMATION"
    c.fill  = hfill(MID_BLUE)
    c.font  = hfont(WHITE, bold=True, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    bi_rows = [
        ("Sprint",        build_info.get("release", sprint)),
        ("Sprint Dates",  build_info.get("sprint_dates", "TBD")),
        ("Build Date",    build_info.get("build_date", "TBD")),
        ("Built By",      build_info.get("built_by", "TBD")),
        ("ELM Sprint",    build_info.get("elm_sprint_url", "")),
    ]
    for i, (label, value) in enumerate(bi_rows, start=3):
        ws.cell(row=i, column=1).value = label
        ws.cell(row=i, column=1).font  = hfont("000000", bold=True, size=10)
        ws.cell(row=i, column=1).fill  = hfill(LIGHT_BLUE)
        ws.cell(row=i, column=1).alignment = Alignment(indent=1)
        ws.merge_cells(start_row=i, end_row=i, start_column=2, end_column=4)
        ws.cell(row=i, column=2).value = value
        ws.cell(row=i, column=2).font  = hfont("000000", size=10)
        ws.cell(row=i, column=2).alignment = Alignment(indent=1, wrap_text=True)
        ws.row_dimensions[i].height = 16

    # ── Spacer row ────────────────────────────────────────────────────────────
    spacer = len(bi_rows) + 3
    ws.row_dimensions[spacer].height = 8

    # ── Environment status header ─────────────────────────────────────────────
    env_start = spacer + 1
    ws.merge_cells(f"A{env_start}:D{env_start}")
    c = ws[f"A{env_start}"]
    c.value = "ENVIRONMENT STATUS"
    c.fill  = hfill(MID_BLUE)
    c.font  = hfont(WHITE, bold=True, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[env_start].height = 18

    env_hdr = env_start + 1
    for col_idx, hdr in enumerate(["Environment", "Status", "Deployed By", "Date", "Notes"], start=1):
        c = ws.cell(row=env_hdr, column=col_idx)
        c.value = hdr
        c.fill  = hfill(ACCENT)
        c.font  = hfont(WHITE, bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[env_hdr].height = 16

    status_fills = {"deployed": GREEN_FILL, "passed": GREEN_FILL, "pending": AMBER_FILL, "failed": RED_FILL}
    for row_offset, env_key in enumerate(["sit", "uat"], start=1):
        r = env_hdr + row_offset
        env = envs.get(env_key, {})
        status = env.get("status", "pending")
        row_fill = hfill(status_fills.get(status.lower(), GREY_FILL))
        values = [
            env_key.upper(),
            status,
            env.get("deployed_by", "—") or "—",
            env.get("deployed_date", "—") or "—",
            env.get("notes", "") or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col_idx)
            c.value = val
            c.fill  = row_fill
            c.font  = hfont("000000", bold=(col_idx == 1), size=10)
            c.alignment = Alignment(horizontal="center" if col_idx <= 4 else "left",
                                    vertical="center", indent=1 if col_idx >= 4 else 0)
        ws.row_dimensions[r].height = 16

    # ── Generated footer ──────────────────────────────────────────────────────
    footer_row = env_hdr + 3
    ws.merge_cells(f"A{footer_row}:D{footer_row}")
    c = ws[f"A{footer_row}"]
    c.value = f"Generated by generate-report.py on {generated_at}"
    c.font  = hfont("808080", size=9)
    c.alignment = Alignment(horizontal="left", indent=1)

    set_col_widths(ws, [22, 18, 18, 42])
    ws.freeze_panes = "A2"


def write_teams_sheet(wb, teams, included_keys, deferred_keys):
    ws = wb.create_sheet("Teams")

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Team", "Source", "Included", "Branch / Folder", "Version / SHA",
               "ELM Work Items", "Summary", "Defer Reason"]
    for col_idx, hdr in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.value = hdr
        c.fill  = hfill(DARK_BLUE)
        c.font  = hfont(WHITE, bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    sc_fills = {"git": GREEN_FILL, "sharepoint": AMBER_FILL}

    for row_idx, key in enumerate(TEAM_ORDER, start=2):
        if key not in teams:
            continue
        t = teams[key]
        label  = TEAM_LABELS.get(key, key)
        sc     = t.get("source_control", "git").lower()
        incl   = t.get("included", False)
        branch = t.get("branch", "")
        version = t.get("commit_or_artifact_version", "")
        url    = t.get("repo_or_sharepoint_url", "")
        elm    = t.get("elm_work_items", "")
        summary = t.get("summary", "")
        reason = t.get("reason_if_deferred", "")

        row_fill = hfill(GREEN_FILL if incl else RED_FILL)
        alt_fill = hfill(sc_fills.get(sc, GREY_FILL))

        row_data = [
            label,
            sc.upper(),
            "YES" if incl else "NO",
            branch if sc == "git" else url,
            version,
            elm,
            summary,
            reason,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value = val
            c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
            c.font = hfont("000000", size=10, bold=(col_idx == 1))
            # Column-specific fills
            if col_idx == 1:
                c.fill = row_fill
            elif col_idx == 2:
                c.fill = alt_fill
            elif col_idx == 3:
                c.fill = hfill(GREEN_FILL if incl else RED_FILL)
                c.font = hfont(GREEN_DARK if incl else "C0392B", bold=True, size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = hfill(GREY_FILL if row_idx % 2 == 0 else WHITE)
            c.border = border_bottom()
        ws.row_dimensions[row_idx].height = 28

    set_col_widths(ws, [26, 12, 10, 32, 18, 28, 38, 38])


def write_cots_sheet(wb, cots_data):
    ws = wb.create_sheet("COTS")

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "COTS Product Details"
    c.fill  = hfill(DARK_BLUE)
    c.font  = hfont(WHITE, bold=True, size=12)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    fields = [
        ("Product",          cots_data.get("product", "")),
        ("Version",          cots_data.get("version", "")),
        ("Previous Version", cots_data.get("previous_version", "")),
        ("Hotfixes",         cots_data.get("hotfixes_included", "")),
        ("Included",         "Yes" if parse_bool(cots_data.get("included", "")) else "No"),
        ("Notes",            cots_data.get("notes", "")),
        ("Release Notes File", cots_data.get("release_notes_file", "")),
    ]
    for i, (label, value) in enumerate(fields, start=2):
        ws.cell(row=i, column=1).value = label
        ws.cell(row=i, column=1).fill  = hfill(LIGHT_BLUE if i % 2 == 0 else WHITE)
        ws.cell(row=i, column=1).font  = hfont("000000", bold=True, size=10)
        ws.cell(row=i, column=1).alignment = Alignment(indent=1, vertical="center")
        ws.cell(row=i, column=2).value = value
        ws.cell(row=i, column=2).fill  = hfill(GREY_FILL if i % 2 == 0 else WHITE)
        ws.cell(row=i, column=2).font  = hfont("000000", size=10)
        ws.cell(row=i, column=2).alignment = Alignment(indent=1, vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 18

    set_col_widths(ws, [22, 48])


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: generate-report.py <sprint>   e.g. generate-report.py sprint-12")

    sprint = sys.argv[1]
    script_dir    = os.path.dirname(os.path.abspath(__file__))
    repo_root     = os.path.dirname(script_dir)
    manifest_path = os.path.join(repo_root, "releases", sprint, "manifest.xlsx")
    output_path   = os.path.join(repo_root, "releases", sprint, "RELEASE-NOTES.xlsx")

    if not os.path.exists(manifest_path):
        sys.exit(f"Error: {manifest_path} not found")

    wb_src = load_workbook(manifest_path, data_only=True)

    build_info = parse_kv_sheet(wb_src["Build Info"])
    teams      = parse_teams_sheet(wb_src["Teams"])
    cots_data  = parse_kv_sheet(wb_src["COTS"])
    envs       = parse_environments_sheet(wb_src["Environments"])

    included_keys = [k for k in TEAM_ORDER if k in teams and teams[k].get("included") is True]
    deferred_keys = [k for k in TEAM_ORDER if k in teams and teams[k].get("included") is not True]

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)  # remove default empty sheet

    write_summary_sheet(wb_out, sprint, build_info, envs, generated_at)
    write_teams_sheet(wb_out, teams, included_keys, deferred_keys)
    write_cots_sheet(wb_out, cots_data)

    wb_out.save(output_path)

    print(f"Report written to: {output_path}")
    print()
    print("=== Next steps: ===")
    print(f"  1. Review {output_path}")
    print(f"  2. git add releases/{sprint}/ && git commit -m 'Release: {sprint} build report'")
    print(f"  3. Open PR titled 'Release {sprint}' for CM team review")
    print(f"  4. Deploy to SIT, then run: ./scripts/tag-environments.sh {sprint} sit")


if __name__ == "__main__":
    main()
