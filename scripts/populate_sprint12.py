#!/usr/bin/env python3
"""
populate_sprint12.py
Opens releases/sprint-12/manifest.xlsx and fills in the sprint-12 data.
Run AFTER create_manifest_xlsx.py sprint-12 has already created the file.

Usage:
    python3 scripts/populate_sprint12.py
"""

import sys
import os

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    sys.exit("Error: openpyxl not found. Install with:  pip3 install openpyxl")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT  = os.path.dirname(SCRIPT_DIR)
XLSX_PATH  = os.path.join(REPO_ROOT, "releases", "sprint-12", "manifest.xlsx")

if not os.path.exists(XLSX_PATH):
    sys.exit(f"Error: {XLSX_PATH} not found. Run create_manifest_xlsx.py sprint-12 first.")

wb = load_workbook(XLSX_PATH)

# ── Sheet 1: Build Info ───────────────────────────────────────────────────────
# Row 1 = header; data rows start at 2
# Row order: release, sprint_dates, build_date, built_by, elm_sprint_url
bi = wb["Build Info"]
bi["B2"] = "sprint-12"
bi["B3"] = "2026-06-09 \u2192 2026-06-20"
bi["B4"] = "2026-06-20"
bi["B5"] = "john.smith"
bi["B6"] = "https://elm.company.com/ccm/web/projects/MyProject#action=com.ibm.team.apt.viewPage&id=sprint-12"

# ── Sheet 2: Teams ────────────────────────────────────────────────────────────
# Columns (1-based):
#  1=team_key  2=team_name  3=source_control  4=repo_or_sharepoint_url
#  5=branch  6=commit_or_artifact_version  7=included  8=elm_work_items
#  9=summary  10=notes  11=depends_on  12=reason_if_deferred

tm = wb["Teams"]

GIT_GREEN  = "D9EAD3"
SP_AMBER   = "FFF2CC"
WARN_AMBER = "FFE699"


def set_team_row(ws, row, team_key, team_name, source_control, url,
                 branch, version, included, elm_items, summary, notes,
                 depends_on, reason):
    ws.cell(row=row, column=1).value  = team_key
    ws.cell(row=row, column=2).value  = team_name
    sc_cell = ws.cell(row=row, column=3)
    sc_cell.value = source_control
    if source_control == "git":
        sc_cell.fill = PatternFill("solid", fgColor=GIT_GREEN)
    elif source_control == "sharepoint":
        sc_cell.fill = PatternFill("solid", fgColor=SP_AMBER)
    ws.cell(row=row, column=4).value  = url
    ws.cell(row=row, column=5).value  = branch
    ws.cell(row=row, column=6).value  = version
    ws.cell(row=row, column=7).value  = included
    ws.cell(row=row, column=8).value  = elm_items
    ws.cell(row=row, column=9).value  = summary
    ws.cell(row=row, column=10).value = notes
    dep_cell = ws.cell(row=row, column=11)
    dep_cell.value = depends_on
    if depends_on:
        dep_cell.fill = PatternFill("solid", fgColor=WARN_AMBER)
    ws.cell(row=row, column=12).value = reason


# Row 2 — Conversion (SharePoint)
set_team_row(
    tm, 2,
    team_key       = "conversion",
    team_name      = "Conversion Team",
    source_control = "sharepoint",
    url            = "https://company.sharepoint.com/sites/ETL/Shared Documents/sprint-12",
    branch         = "N/A",
    version        = "v1.3",
    included       = "TRUE",
    elm_items      = "ELM-1201 ELM-1202 ELM-1208",
    summary        = "Customer address mapping null-safe account load",
    notes          = "",
    depends_on     = "",
    reason         = "",
)

# Row 3 — Interfaces (git) — depends on func_config
set_team_row(
    tm, 3,
    team_key       = "interfaces",
    team_name      = "Interfaces Team",
    source_control = "git",
    url            = "satishreddy13/repo-interfaces",
    branch         = "sprint/sprint-12",
    version        = "b2c3d4e5",
    included       = "TRUE",
    elm_items      = "ELM-1210 ELM-1215",
    summary        = "REST adapter v2 endpoint updates retry logic",
    notes          = "",
    depends_on     = "func_config",
    reason         = "",
)

# Row 4 — Workflow Config (git, deferred)
set_team_row(
    tm, 4,
    team_key       = "workflow_config",
    team_name      = "Workflow Config Team",
    source_control = "git",
    url            = "satishreddy13/repo-workflow-config",
    branch         = "sprint/sprint-12",
    version        = "",
    included       = "FALSE",
    elm_items      = "ELM-1220 ELM-1221",
    summary        = "",
    notes          = "",
    depends_on     = "",
    reason         = "ELM-1220 blocked pending sign-off from business deferred to sprint-13",
)

# Row 5 — Functional Product Config (git)
set_team_row(
    tm, 5,
    team_key       = "func_config",
    team_name      = "Functional Product Config",
    source_control = "git",
    url            = "satishreddy13/repo-func-config",
    branch         = "sprint/sprint-12",
    version        = "c3d4e5f6",
    included       = "TRUE",
    elm_items      = "ELM-1230 ELM-1231",
    summary        = "Product configuration tables updated for region codes",
    notes          = "",
    depends_on     = "",
    reason         = "",
)

# ── Sheet 3: COTS ─────────────────────────────────────────────────────────────
# Row order: product, version, previous_version, hotfixes_included, included, notes, release_notes_file
ct = wb["COTS"]
ct["B2"] = "Acme ETL Platform"
ct["B3"] = "2.4.0"
ct["B4"] = "2.3.8"
ct["B5"] = "HF-2024-011 HF-2024-012"
ct["B6"] = "TRUE"
ct["B7"] = "Upgrade from 2.3.8 schema change in audit_log table"
ct["B8"] = "cots-sprint-12.md"

# ── Sheet 4: Environments ─────────────────────────────────────────────────────
# Columns: environment, status, deployed_by, deployed_date, passed_date, notes
en = wb["Environments"]

en["A2"] = "sit"
en["B2"] = "deployed"
en["C2"] = "john.smith"
en["D2"] = "2026-06-20 14:30"
en["E2"] = ""
en["F2"] = ""

en["A3"] = "uat"
en["B3"] = "pending"
en["C3"] = ""
en["D3"] = ""
en["E3"] = ""
en["F3"] = "UAT currently running sprint-11"

wb.save(XLSX_PATH)
print(f"Populated: {XLSX_PATH}")
